#!/usr/bin/python
#-*-coding:UTF-8-*-
# encoding=utf8
from com.office.util.fileUtil import getFiles
def read(filename, sheetName='Sheet1'):
    import xlrd
    data = xlrd.open_workbook(filename);
    table = data.sheet_by_name(sheetName)
    nrows = table.nrows
    returnList = [];
    for i in range(nrows ):
        returnList.append(table.row_values(i))
    return returnList;

def write(tableData, filename, sheetName='Sheet1'):
    import xlwt
    wb=xlwt.Workbook()
    ws=wb.add_sheet(sheetName)
    for i in range(len(tableData)):
        rowDatas = tableData[i];
        for j in range(len(rowDatas)):
            ws.write(i,j,rowDatas[j])
    wb.save(filename)
    
def writeExcel(tableData, filename, sheetName='Sheet1'):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.create_sheet(sheetName, 0)
    
    for i in range(len(tableData)):
        rowDatas = tableData[i];
        if type(rowDatas) != list:
            rowDatas = [rowDatas];
        ws.append(rowDatas)
    # Save the file
    wb.save(filename)
    
if __name__ == "__main__":
    ownerMap = {};
    relationMap = {}
    table = read(r"/Users/jyang/Desktop/责任表/脚本源.xlsx", sheetName=u"工作表3")
    #清洗
    newtable = [];
    for line in table:
        newLine = [];
        for i in line:
            newLine.append(i.lower().strip().replace(",","\n"));
        newtable.append(newLine)
    table = newtable
#     print "len", len(table)
    allDbTables = [];#全部表名
    putMap = {}#表名：使用者
    #初始化
    for line in table:
        putName = line[0].strip();
        if not putName:
            continue
        if len(line) >= 2 and line[1]:
            tableNm = line[1].strip();
            allDbTables.append(tableNm)
            putMap[tableNm] = putName
        if len(line) >= 3 and line[2]:
            for sp in line[2].replace(";","\n").replace(u"，","\n").split("\n"):
                tableNm = sp.strip();
                allDbTables.append(tableNm)
                putMap[tableNm] = putName
#     print "lenputMap", len(putMap)
    allDbTables = set(allDbTables)
    print "总共涉及%d张表"%len(allDbTables)
    #process
    for line in table:
        try:
            tableNm = line[1].strip()
            if len(line) >= 2 and line[0]:
                ownerMap[line[1]] = line[0];
            if len(line) >=3 and line[2]:
                relationMap[line[1]] = line[2];
        except Exception,e:
            print e
            print line
                
                
    def printNoOwner():
        count = 0
        for tbNm in sorted(allDbTables):
            if tbNm and not ownerMap.has_key(tbNm):
                print tbNm,",",putMap[tbNm]
                count += 1
        print "还有%d张表没有责任人，表后面跟的名字是表名的提供者。"%count
#     printNoOwner()
                
    def printNoParent():
        count = 0
        for tbNm in sorted(allDbTables):
            if tbNm and not relationMap.has_key(tbNm):
                print tbNm,",",putMap[tbNm]
                count += 1
        print "还有%d张表没有上游，需要确认是否是源头了"%count
#     
    def writeResult():
        tableData = [[u"表名",u"责任人",u"上游表"]]
        for t in sorted(allDbTables):
            if not t:
                continue
            owner = u"无"
            table = t
            parent = ""
            if ownerMap.has_key(t):
                owner = ownerMap[t];

            if relationMap.has_key(t):
                parent = relationMap[t]
            owner = owner.replace(u"无", u"未知").replace(u"未知", u"待确认")
            tableData.append([table,owner,parent.replace(";","\n").replace(u"，","\n")])
        write(tableData, "/Users/jyang/Desktop/主城区城市大脑表依赖关系和责任人.xls", "sheet1")

    writeResult()
#     printNoParent()
    printNoOwner();
    print "总共涉及%d张表"%len(allDbTables)
        
